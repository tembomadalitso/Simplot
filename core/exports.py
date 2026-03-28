import io
from datetime import date
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from .models import Property, RentalAgreement, Expense

User = get_user_model()

def export_xlsx_view(request):
    """Generates an XLSX report for properties/income or compliance with filtering."""
    user = request.user
    if not user.is_authenticated:
        return HttpResponse("Unauthorized", status=401)

    if user.user_type not in ['LANDLORD', 'ZRA', 'MINISTRY']:
        return HttpResponseForbidden("You do not have permission to export reports.")

    wb = Workbook()
    ws = wb.active

    if user.user_type == 'LANDLORD':
        ws.title = "Portfolio Report"
        search = request.GET.get('search', '')
        category = request.GET.get('category', '')
        properties = Property.objects.filter(owner=user)
        if search: properties = properties.filter(title__icontains=search)
        if category: properties = properties.filter(category=category)

        headers = ["Listing ID", "Property Title", "Category", "District", "Price (ZMW)", "Compliance"]
        ws.append(headers)
        for p in properties:
            ws.append([p.id, p.title, p.get_category_display(), p.district, float(p.price), "Verified" if p.is_tax_compliant else "Pending"])

    elif user.user_type == 'ZRA':
        ws.title = "Tax Compliance Audit"
        search_name = request.GET.get('name', '')
        status_filter = request.GET.get('status', '') # 'true' or 'false'

        landlords = User.objects.filter(user_type='LANDLORD')
        if search_name:
            landlords = landlords.filter(username__icontains=search_name) | landlords.filter(nrc_number__icontains=search_name)

        headers = ["Landlord", "Email", "TPIN", "Property Count", "Est. Annual Revenue"]
        ws.append(headers)
        for l in landlords:
            props = Property.objects.filter(owner=l)
            if status_filter == 'true':
                props = props.filter(is_tax_compliant=True)
            elif status_filter == 'false':
                props = props.filter(is_tax_compliant=False)

            if props.exists() or not (status_filter):
                ws.append([l.username, l.email, l.tpin_number or 'N/A', props.count(), sum(float(p.price) * 12 for p in props)])

    elif user.user_type == 'MINISTRY':
        ws.title = "District Occupancy"
        district = request.GET.get('district', '')
        category = request.GET.get('category', '')

        agreements = RentalAgreement.objects.filter(is_active=True)
        if district: agreements = agreements.filter(property__district__icontains=district)
        if category: agreements = agreements.filter(property__category=category)

        headers = ["District", "Property", "Category", "Occupants", "Status"]
        ws.append(headers)
        for ag in agreements:
            ws.append([ag.property.district, ag.property.title, ag.property.get_category_display(), ag.number_of_occupants, ag.status])

    # Header Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Report_{user.username}.xlsx"'
    return response

def export_pdf_report(request):
    """Generates a PDF summary report for all roles with filtering."""
    user = request.user
    if not user.is_authenticated:
        return HttpResponse("Unauthorized", status=401)

    if user.user_type not in ['LANDLORD', 'ZRA', 'MINISTRY']:
        return HttpResponseForbidden("You do not have permission to export reports.")

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 20)
    p.drawString(100, height - 80, "ZAMBIA RENTALS — PREMIUM REPORT")
    p.setFont("Helvetica", 10)
    p.drawString(100, height - 100, f"Role: {user.get_user_type_display()} | Generated: {date.today()}")
    p.line(100, height - 110, 500, height - 110)

    y = height - 140
    if user.user_type == 'LANDLORD':
        search = request.GET.get('search', '')
        category = request.GET.get('category', '')
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y, f"Financial Summary & Portfolio (Filter: {search or 'None'}, {category or 'All'})")
        y -= 30
        props = Property.objects.filter(owner=user)
        if search: props = props.filter(title__icontains=search)
        if category: props = props.filter(category=category)

        for prop in props:
            p.setFont("Helvetica", 10)
            p.drawString(100, y, f"• {prop.title} - K{prop.price}/mo - {prop.district}")
            y -= 15
            if y < 100: p.showPage(); y = height - 100

    elif user.user_type == 'ZRA':
        search_name = request.GET.get('name', '')
        status_filter = request.GET.get('status', '')
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y, "Tax Compliance & Revenue Audit")
        y -= 30
        all_props = Property.objects.all()
        if search_name:
            all_props = all_props.filter(owner__username__icontains=search_name) | all_props.filter(owner__nrc_number__icontains=search_name)
        if status_filter == 'true': all_props = all_props.filter(is_tax_compliant=True)
        elif status_filter == 'false': all_props = all_props.filter(is_tax_compliant=False)

        for prop in all_props:
            p.setFont("Helvetica", 10)
            p.drawString(100, y, f"• {prop.title} | Owner: {prop.owner.username} | Compliant: {prop.is_tax_compliant}")
            y -= 15
            if y < 100: p.showPage(); y = height - 100

    elif user.user_type == 'MINISTRY':
        district = request.GET.get('district', '')
        category = request.GET.get('category', '')
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y, f"National Occupancy & Density Audit (Filter: {district or 'All Districts'})")
        y -= 30
        agreements = RentalAgreement.objects.filter(is_active=True)
        if district: agreements = agreements.filter(property__district__icontains=district)
        if category: agreements = agreements.filter(property__category=category)

        for ag in agreements:
            p.setFont("Helvetica", 10)
            p.drawString(100, y, f"• {ag.property.district} | {ag.property.title} | Residents: {ag.number_of_occupants}")
            y -= 15
            if y < 100: p.showPage(); y = height - 100

    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Report_{user.username}.pdf"'
    return response
